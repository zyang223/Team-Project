#DCF template created by Zide Yang (April 13 2023)
#pip install openpyxl
Api key="3178e48c4517a7c30aa49ccbadef2582"
from openpyxl import Workbook, load_workbook
#create a work book
# workbook=Workbook()
# worksheet=workbook.active
# worksheet.title="DCF Template"

workbook=load_workbook("DCF Template.xlsx")

worksheet=workbook.active

headers=["Name","Ticker","Date"]
worksheet.append(headers)

share_price=["999.99"]
worksheet.append(share_price)


years=["2018"]
worksheet.append(years)

#save the work book
workbook.save("DCF Template.xlsx")  
